#!/usr/bin/env python3
"""Import secretary Excel workbooks into clean weekly + special_events CSVs."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit(
        "openpyxl is required. Install with: python3 -m pip install --user openpyxl"
    ) from e

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

DEFAULT_WORSHIP = Path(
    "/Users/paul/Library/CloudStorage/Dropbox-Personal/00 CONCORD UMC/"
    "ATTENDANCE PROJECT DATA/RAW DATA/Attendance - Worship.xlsx"
)
DEFAULT_XMAS = Path(
    "/Users/paul/Library/CloudStorage/Dropbox-Personal/00 CONCORD UMC/"
    "ATTENDANCE PROJECT DATA/RAW DATA/Christmas Eve Attendance.xlsx"
)

# Known Easter Sundays (Western)
EASTER = {
    date(2021, 4, 4),
    date(2022, 4, 17),
    date(2023, 4, 9),
    date(2024, 3, 31),
    date(2025, 4, 20),
    date(2026, 4, 5),
}

# Header typos / bad strings → correct date
DATE_FIXES = {
    "1/18/25/26": date(2026, 1, 18),
    "02/23/205": date(2025, 2, 23),
    "0302/2025": date(2025, 3, 2),
}

WEEKLY_FIELDS = [
    "date",
    "year",
    "month",
    "trad_9",
    "cont_9",
    "kids_11",
    "trad_11",
    "cont_11",
    "boxcast_9",
    "youtube_9",
    "facebook_9",
    "boxcast_11",
    "youtube_11",
    "facebook_11",
    "in_person",
    "online",
    "total",
    "is_snow",
    "is_easter",
    "is_christmas_eve",
    "exclude_from_averages",
    "notes",
]

SPECIAL_FIELDS = [
    "date",
    "year",
    "event_type",
    "service_label",
    "in_person",
    "online",
    "boxcast",
    "youtube",
    "facebook",
    "notes",
]


def parse_date(value, sheet_year: int | None = None) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        d = value.date()
        return d if d.year >= 2018 else None
    if isinstance(value, date):
        return value if value.year >= 2018 else None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel serials for 2018–2035 are ~43101–50754. Do NOT treat
        # ordinary attendance counts (typically < 2000) as dates.
        serial = float(value)
        if serial < 43000 or serial > 55000:
            return None
        try:
            d = (datetime(1899, 12, 30) + timedelta(days=int(serial))).date()
            return d if d.year >= 2018 else None
        except Exception:
            return None
    s = str(value).strip()
    if s in DATE_FIXES:
        return DATE_FIXES[s]
    # stray "c" and similar junk
    if len(s) <= 2 and not s[0].isdigit():
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m/%d/%Y %H:%M:%S"):
        try:
            d = datetime.strptime(s.split()[0], fmt).date()
            return d if d.year >= 2018 else None
        except ValueError:
            pass
    # 0302/2025 style already handled; try mmdd/yyyy
    m = re.match(r"^(\d{1,2})(\d{2})/(\d{4})$", s)
    if m:
        d = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        return d if d.year >= 2018 else None
    return None


def to_num(value):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        t = value.strip()
        if t.upper() in {"SNOW", "NA", "N/A", "N\\A"}:
            return t.upper() if t.upper() == "SNOW" else None
        if t.upper() == "SNOW":
            return "SNOW"
        try:
            return int(float(t.replace(",", "")))
        except ValueError:
            return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    return None


def find_date_row(ws, max_scan=6):
    best = None
    best_count = 0
    for r in range(1, max_scan + 1):
        dates = {}
        for c in range(1, min(ws.max_column or 1, 90) + 1):
            d = parse_date(ws.cell(r, c).value)
            if d:
                dates[c] = d
        if len(dates) > best_count:
            best_count = len(dates)
            best = (r, dates)
    return best


def map_labels(ws):
    """Map normalized label keywords → row index."""
    rows = {}
    for r in range(1, (ws.max_row or 0) + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str):
            continue
        key = label.strip().lower().replace("  ", " ")
        rows[key] = r
    return rows


def pick_row(label_map, *candidates):
    for cand in candidates:
        for key, r in label_map.items():
            if cand in key:
                return r
    return None


def empty_week(d: date, notes: str = "") -> dict:
    is_easter = d in EASTER
    is_xmas_eve = d.month == 12 and d.day == 24
    is_xmas_day = d.month == 12 and d.day == 25
    return {
        "date": d.isoformat(),
        "year": d.year,
        "month": d.month,
        "trad_9": None,
        "cont_9": None,
        "kids_11": None,
        "trad_11": None,
        "cont_11": None,
        "boxcast_9": None,
        "youtube_9": None,
        "facebook_9": None,
        "boxcast_11": None,
        "youtube_11": None,
        "facebook_11": None,
        "in_person": None,
        "online": None,
        "total": None,
        "is_snow": False,
        "is_easter": is_easter,
        "is_christmas_eve": is_xmas_eve,
        "exclude_from_averages": average_exclusion_reason(d, snow=False, in_person=None) is not None,
        "notes": notes,
    }


def average_exclusion_reason(d: date, snow: bool = False, in_person=None) -> str | None:
    """Weekly averages exclude: full snow closures, Christmas Eve, Christmas Day on Sunday.
    Easter and the Sunday after Christmas stay in averages.
    """
    if snow and (in_person is None or in_person == 0):
        return "snow closure"
    if d.month == 12 and d.day == 24:
        return "Christmas Eve"
    if d.month == 12 and d.day == 25 and d.weekday() == 6:  # Sunday
        return "Christmas Day on Sunday"
    return None


def finalize_week(row: dict) -> dict:
    snow = row["is_snow"]
    nums = {
        k: row[k]
        for k in (
            "trad_9",
            "cont_9",
            "kids_11",
            "trad_11",
            "cont_11",
            "boxcast_9",
            "youtube_9",
            "facebook_9",
            "boxcast_11",
            "youtube_11",
            "facebook_11",
        )
    }
    # Detect snow if any in-person field was SNOW marker earlier
    in_person_parts = []
    for k in ("trad_9", "cont_9", "kids_11", "trad_11", "cont_11"):
        v = nums[k]
        if v == "SNOW":
            snow = True
            row[k] = None
        elif isinstance(v, int):
            in_person_parts.append(v)
            row[k] = v
        else:
            row[k] = None if v is None else v

    online_parts = []
    for k in (
        "boxcast_9",
        "youtube_9",
        "facebook_9",
        "boxcast_11",
        "youtube_11",
        "facebook_11",
    ):
        v = nums[k]
        if v == "SNOW":
            snow = True
            row[k] = None
        elif isinstance(v, int):
            online_parts.append(v)
            row[k] = v
        else:
            row[k] = None

    row["is_snow"] = bool(snow)
    d = date.fromisoformat(row["date"])
    row["is_easter"] = d in EASTER
    row["is_christmas_eve"] = d.month == 12 and d.day == 24

    if snow and not in_person_parts:
        # Fully closed snow Sunday
        row["in_person"] = 0
        row["online"] = sum(online_parts) if online_parts else None
        row["total"] = row["online"] or 0
        if "snow" not in (row["notes"] or "").lower():
            row["notes"] = (row["notes"] + "; snow closure").strip("; ")
    elif snow and sum(in_person_parts) == 0:
        # Snow flagged and only zeros recorded for in-person
        row["in_person"] = 0
        row["online"] = sum(online_parts) if online_parts else None
        row["total"] = row["online"] or 0
        if "snow" not in (row["notes"] or "").lower():
            row["notes"] = (row["notes"] + "; snow closure").strip("; ")
    else:
        row["in_person"] = sum(in_person_parts) if in_person_parts else None
        row["online"] = sum(online_parts) if online_parts else None
        if row["in_person"] is not None or row["online"] is not None:
            row["total"] = (row["in_person"] or 0) + (row["online"] or 0)
        else:
            row["total"] = None
        if snow:
            # Limited weather Sunday with some in-person counts — keep in averages
            if "weather" not in (row["notes"] or "").lower():
                row["notes"] = (row["notes"] + "; limited weather Sunday").strip("; ")

    if row["is_easter"] and "easter" not in (row["notes"] or "").lower():
        row["notes"] = (row["notes"] + "; Easter Sunday").strip("; ")

    reason = average_exclusion_reason(
        d, snow=row["is_snow"], in_person=row.get("in_person")
    )
    row["exclude_from_averages"] = reason is not None
    if reason and reason not in (row["notes"] or "").lower():
        if reason == "Christmas Eve" and "christmas eve" not in (row["notes"] or "").lower():
            row["notes"] = (row["notes"] + "; Christmas Eve").strip("; ")
        elif reason == "Christmas Day on Sunday":
            row["notes"] = (row["notes"] + "; Christmas Day on Sunday").strip("; ")
    return row


def import_year_sheet(ws, sheet_name: str) -> list[dict]:
    found = find_date_row(ws)
    if not found:
        print(f"  skip {sheet_name}: no date row")
        return []
    date_row, dates = found
    label_map = map_labels(ws)

    r_trad9 = pick_row(label_map, "1st service - traditional")
    r_cont9 = pick_row(label_map, "1st service - contemporary")
    r_box9 = pick_row(label_map, "1st service - boxcast")
    r_yt9 = pick_row(label_map, "1st service - youtube", "1st service - youtube")
    r_fb9 = pick_row(label_map, "1st service -facebook", "1st service - facebook")
    r_trad11 = pick_row(label_map, "2nd service - traditional")
    r_cont11 = pick_row(label_map, "2nd service - contemporary")
    r_kids = pick_row(label_map, "kids worship")
    r_box11 = pick_row(label_map, "2nd service - boxcast")
    r_yt11 = pick_row(label_map, "2nd service - youtube")
    r_fb11 = pick_row(label_map, "2nd service - facebook")

    # Fix known bad date: 2021 sheet has one column mistyped as 2023-05-30
    # which should be 2021-05-30 (between May 23 and June 6)
    fixed_dates = {}
    sorted_cols = sorted(dates.items())
    for i, (c, d) in enumerate(sorted_cols):
        if d.year == 2023 and "2021" in sheet_name and d.month == 5 and d.day == 30:
            d = date(2021, 5, 30)
        # Sequence repair: if date is wildly out of order vs neighbors, try same month/day in sheet year
        fixed_dates[c] = d

    # Also repair Oct 12 2025 if header was "c" — may be missing entirely
    weeks = {}
    for c, d in fixed_dates.items():
        # Drop clearly wrong years for named sheets
        m = re.search(r"(20\d{2})", sheet_name)
        if m:
            expect = int(m.group(1))
            if d.year not in {expect - 1, expect, expect + 1}:
                continue

        mapping = {
            "trad_9": r_trad9,
            "cont_9": r_cont9,
            "boxcast_9": r_box9,
            "youtube_9": r_yt9,
            "facebook_9": r_fb9,
            "trad_11": r_trad11,
            "cont_11": r_cont11,
            "kids_11": r_kids,
            "boxcast_11": r_box11,
            "youtube_11": r_yt11,
            "facebook_11": r_fb11,
        }
        raw_vals = {}
        for field, r in mapping.items():
            if r is None:
                continue
            raw_vals[field] = to_num(ws.cell(r, c).value)

        # Skip future / empty date columns (headers present, no counts yet)
        has_any = any(v is not None for v in raw_vals.values())
        if not has_any:
            continue

        row = empty_week(d)
        row.update(raw_vals)
        weeks[d.isoformat()] = finalize_week(row)

    print(f"  {sheet_name}: {len(weeks)} Sundays")
    return list(weeks.values())


def import_jan_feb_2021(ws) -> list[dict]:
    """COVID-era layout: dates in column A, channels across columns."""
    weeks = []
    # Columns from earlier inspection
    # B 8:45 Boxcast, C 8:45 FB, D 8:45 church, E 10 Boxcast, F 10 FB,
    # G 11 Boxcast, H 11 FB, I 10 church, ... TOTAL
    for r in range(3, 12):
        d = parse_date(ws.cell(r, 1).value)
        if not d:
            continue
        row = empty_week(d, notes="COVID-era Jan–Feb 2021 schedule")
        church_845 = to_num(ws.cell(r, 4).value)  # often N/A
        church_10 = to_num(ws.cell(r, 9).value)
        box_845 = to_num(ws.cell(r, 2).value)
        fb_845 = to_num(ws.cell(r, 3).value)
        box_10 = to_num(ws.cell(r, 5).value)
        fb_10 = to_num(ws.cell(r, 6).value)
        box_11 = to_num(ws.cell(r, 7).value)
        fb_11 = to_num(ws.cell(r, 8).value)

        # Map roughly into trad/cont buckets is ambiguous; store online + in_person totals
        in_person = 0
        for v in (church_845, church_10):
            if isinstance(v, int):
                in_person += v
        online = 0
        for v in (box_845, fb_845, box_10, fb_10, box_11, fb_11):
            if isinstance(v, int):
                online += v
        # evening streams
        for c in range(10, 14):
            v = to_num(ws.cell(r, c).value)
            if isinstance(v, int):
                online += v

        row["in_person"] = in_person if in_person else None
        row["online"] = online if online else None
        row["boxcast_9"] = box_845 if isinstance(box_845, int) else None
        row["facebook_9"] = fb_845 if isinstance(fb_845, int) else None
        row["boxcast_11"] = box_11 if isinstance(box_11, int) else None
        row["facebook_11"] = fb_11 if isinstance(fb_11, int) else None
        # Put 10 AM church under trad_11 as best available in-person slot
        row["trad_11"] = church_10 if isinstance(church_10, int) else None
        if row["in_person"] is not None or row["online"] is not None:
            row["total"] = (row["in_person"] or 0) + (row["online"] or 0)
        row["exclude_from_averages"] = False
        weeks.append(row)
    print(f"  2021 Attendance_Jan & Feb: {len(weeks)} Sundays")
    return weeks


def import_christmas_eve_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    events = []

    # Year headers on row 3: YEAR, 2025, blank, blank, 2024, ...
    year_cols = {}
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(3, c).value
        if isinstance(v, (int, float)) and 2015 <= int(v) <= 2030:
            year_cols[int(v)] = c

    # Totals row 5
    totals = {}
    for year, c in year_cols.items():
        t = to_num(ws.cell(5, c).value)
        if isinstance(t, int):
            totals[year] = t

    # Service blocks: in-person roughly rows 7-12, youtube 16+, boxcast 24+, facebook 32+
    # Parse label+value pairs in each year column cluster (label at c-0 or c for some)
    def services_for_year(year: int, value_col: int, start_row: int, end_row: int, channel: str):
        out = []
        for r in range(start_row, end_row + 1):
            # label is usually in the same column group: for 2025 labels in col A (1), values in col B (2)
            # Structure: [label, value, None, label, value, None, ...]
            label_col = value_col - 1
            label = ws.cell(r, label_col).value
            val = to_num(ws.cell(r, value_col).value)
            if not isinstance(label, str) or not isinstance(val, int):
                continue
            lab = label.strip()
            if not lab or lab.upper() in {"YOUTUBE", "BOXCAST", "FACEBOOK", "TOTALS", "YEAR"}:
                continue
            out.append((lab, val))
        return out

    for year, c in sorted(year_cols.items(), reverse=True):
        d = date(year, 12, 24)
        # In person services
        for lab, val in services_for_year(year, c, 7, 12, "in_person"):
            events.append(
                {
                    "date": d.isoformat(),
                    "year": year,
                    "event_type": "christmas_eve",
                    "service_label": lab,
                    "in_person": val,
                    "online": None,
                    "boxcast": None,
                    "youtube": None,
                    "facebook": None,
                    "notes": "",
                }
            )
        for lab, val in services_for_year(year, c, 15, 21, "youtube"):
            events.append(
                {
                    "date": d.isoformat(),
                    "year": year,
                    "event_type": "christmas_eve",
                    "service_label": lab,
                    "in_person": None,
                    "online": val,
                    "boxcast": None,
                    "youtube": val,
                    "facebook": None,
                    "notes": "youtube",
                }
            )
        for lab, val in services_for_year(year, c, 23, 29, "boxcast"):
            events.append(
                {
                    "date": d.isoformat(),
                    "year": year,
                    "event_type": "christmas_eve",
                    "service_label": lab,
                    "in_person": None,
                    "online": val,
                    "boxcast": val,
                    "youtube": None,
                    "facebook": None,
                    "notes": "boxcast",
                }
            )
        for lab, val in services_for_year(year, c, 31, 37, "facebook"):
            events.append(
                {
                    "date": d.isoformat(),
                    "year": year,
                    "event_type": "christmas_eve",
                    "service_label": lab,
                    "in_person": None,
                    "online": val,
                    "boxcast": None,
                    "youtube": None,
                    "facebook": val,
                    "notes": "facebook",
                }
            )
        if year in totals:
            events.append(
                {
                    "date": d.isoformat(),
                    "year": year,
                    "event_type": "christmas_eve",
                    "service_label": "TOTAL_IN_PERSON",
                    "in_person": totals[year],
                    "online": None,
                    "boxcast": None,
                    "youtube": None,
                    "facebook": None,
                    "notes": "sheet total row",
                }
            )

    print(f"  Christmas Eve Attendance.xlsx: {len(events)} rows")
    return events


def import_easter_from_weekly(weeks: list[dict]) -> list[dict]:
    events = []
    by_date = {w["date"]: w for w in weeks}
    for d in sorted(EASTER):
        key = d.isoformat()
        w = by_date.get(key)
        if not w:
            continue
        events.append(
            {
                "date": key,
                "year": d.year,
                "event_type": "easter",
                "service_label": "TOTAL_IN_PERSON",
                "in_person": w.get("in_person"),
                "online": w.get("online"),
                "boxcast": None,
                "youtube": None,
                "facebook": None,
                "notes": "from weekly sheet",
            }
        )
        for lab, field in (
            ("9AM Traditional", "trad_9"),
            ("9AM Contemporary", "cont_9"),
            ("11AM Traditional", "trad_11"),
            ("11AM Contemporary", "cont_11"),
            ("Kids Worship 11AM", "kids_11"),
        ):
            if w.get(field) is not None:
                events.append(
                    {
                        "date": key,
                        "year": d.year,
                        "event_type": "easter",
                        "service_label": lab,
                        "in_person": w[field],
                        "online": None,
                        "boxcast": None,
                        "youtube": None,
                        "facebook": None,
                        "notes": "",
                    }
                )
    return events


def import_holy_week(ws) -> list[dict]:
    events = []
    current_year = None
    for r in range(1, (ws.max_row or 0) + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and "HOLY WEEK" in a.upper():
            m = re.search(r"(20\d{2})", a)
            if m:
                current_year = int(m.group(1))
            continue
        if isinstance(a, str) and "ASH WEDNESDAY" in a.upper():
            m = re.search(r"(20\d{2})", a)
            y = int(m.group(1)) if m else current_year
            if y:
                events.append(
                    {
                        "date": f"{y}-02-01",  # approximate placeholder if unknown
                        "year": y,
                        "event_type": "ash_wednesday",
                        "service_label": str(ws.cell(r, 2).value or "Ash Wednesday"),
                        "in_person": to_num(ws.cell(r, 3).value)
                        if isinstance(to_num(ws.cell(r, 3).value), int)
                        else to_num(ws.cell(r, 2).value),
                        "online": None,
                        "boxcast": None,
                        "youtube": None,
                        "facebook": None,
                        "notes": a,
                    }
                )
            continue
        if not current_year or not isinstance(a, str) or not a.strip():
            continue
        lab = a.strip()
        if lab.upper().startswith("HOLY"):
            continue
        ip = to_num(ws.cell(r, 3).value)
        yt = to_num(ws.cell(r, 4).value)
        box = to_num(ws.cell(r, 5).value)
        if not any(isinstance(x, int) for x in (ip, yt, box)):
            continue
        # Approximate dates within Holy Week — use year only for grouping
        events.append(
            {
                "date": f"{current_year}-03-01",
                "year": current_year,
                "event_type": "holy_week",
                "service_label": lab,
                "in_person": ip if isinstance(ip, int) else None,
                "online": (yt if isinstance(yt, int) else 0)
                + (box if isinstance(box, int) else 0)
                or None,
                "boxcast": box if isinstance(box, int) else None,
                "youtube": yt if isinstance(yt, int) else None,
                "facebook": None,
                "notes": "",
            }
        )
    print(f"  Lent_Holy Week: {len(events)} rows")
    return events


def write_csv(path: Path, fields: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in rows:
            out = {}
            for k in fields:
                v = row.get(k)
                if isinstance(v, bool):
                    out[k] = "true" if v else "false"
                elif v is None:
                    out[k] = ""
                else:
                    out[k] = v
            w.writerow(out)


def main():
    worship = DEFAULT_WORSHIP
    xmas = DEFAULT_XMAS
    if not worship.exists():
        raise SystemExit(f"Missing worship workbook: {worship}")

    print(f"Loading {worship.name}…")
    wb = openpyxl.load_workbook(worship, data_only=True)

    all_weeks: dict[str, dict] = {}

    # Jan–Feb 2021 first
    if "2021 Attendance_Jan & Feb" in wb.sheetnames:
        for w in import_jan_feb_2021(wb["2021 Attendance_Jan & Feb"]):
            all_weeks[w["date"]] = w

    year_sheets = [
        "2021 Attendance",
        "2022 Attendance",
        "2023 Attendance",
        "2024 Attendance",
        "2025 Attendance",
        "2026 Attendance",
    ]
    for name in year_sheets:
        if name not in wb.sheetnames:
            continue
        for w in import_year_sheet(wb[name], name):
            # Later sheets / fuller grids overwrite earlier when same date
            prev = all_weeks.get(w["date"])
            if prev and prev.get("notes") and not w.get("notes"):
                w["notes"] = prev["notes"]
            all_weeks[w["date"]] = w

    # Easter 2022 column is blank in the weekly sheet — backfill from prior cleaned totals
    easter_2022 = date(2022, 4, 17).isoformat()
    if easter_2022 not in all_weeks:
        row = empty_week(date(2022, 4, 17), notes="Easter 2022 backfilled; weekly sheet column blank")
        row["trad_9"] = None
        row["cont_9"] = None
        row["trad_11"] = 570  # combined traditional for three-service morning
        row["cont_11"] = 711  # combined contemporary
        row["kids_11"] = None
        row["in_person"] = 1281
        row["online"] = None
        row["total"] = 1281
        row["is_easter"] = True
        row["exclude_from_averages"] = False  # Easter stays in weekly averages
        all_weeks[easter_2022] = row
        print("  backfilled Easter 2022 (2022-04-17)")

    # Flag full snow closures more carefully: if is_snow and in_person==0
    weeks = [all_weeks[k] for k in sorted(all_weeks)]
    print(f"Weekly Sundays total: {len(weeks)}")

    specials: list[dict] = []
    if xmas.exists():
        print(f"Loading {xmas.name}…")
        specials.extend(import_christmas_eve_xlsx(xmas))
    else:
        print("Christmas Eve workbook not found — skipping")

    specials.extend(import_easter_from_weekly(weeks))
    if "Lent_Holy Week" in wb.sheetnames:
        specials.extend(import_holy_week(wb["Lent_Holy Week"]))

    write_csv(DATA / "weekly_attendance.csv", WEEKLY_FIELDS, weeks)
    write_csv(DATA / "special_events.csv", SPECIAL_FIELDS, specials)
    print(f"Wrote {DATA / 'weekly_attendance.csv'}")
    print(f"Wrote {DATA / 'special_events.csv'} ({len(specials)} rows)")


if __name__ == "__main__":
    main()
